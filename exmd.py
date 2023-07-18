import re
import os
import openpyxl
import warnings
import logging
import datetime
import argparse

# Global hardcoded values

# колонка С в вхідному exel документі
key_name_coulmn_index = 2

# розмірність (кількість) штю, кг, г ...
unit_coulmn_index = 3

# заказ
order_coulmn_index = 4

sheet_name = "Лист1"

# Create a custom UnicodeStreamHandler with 'utf-8' encoding
class UnicodeStreamHandler(logging.StreamHandler):
    def emit(self, record):
        try:
            msg = self.format(record)
            stream = self.stream
            fs = "%s\n"
            # Note: Using 'utf-8' encoding for writing log messages
            if hasattr(stream, 'encoding') and stream.encoding not in (None, 'utf-8'):
                fs = fs.encode('utf-8')
            stream.write(fs % msg)
            self.flush()
        except Exception:
            self.handleError(record)

# Create a custom UnicodeFileHandler with 'utf-8' encoding
class UnicodeFileHandler(logging.FileHandler):
    def __init__(self, filename, mode='a', encoding=None, delay=False):
        if encoding is None:
            encoding = 'utf-8'
        super().__init__(filename, mode, encoding, delay)

def setup_logging(log_file="log.txt"):
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            UnicodeStreamHandler(),                 # For printing to the console
            # For writing to the log.txt file
            UnicodeFileHandler(log_file, mode="w"),
        ]
    )

def fix_unit(input_str):
    if not input_str or input_str == "банка":
        return "шт"
    else:
        return input_str

def strip_key_val(val):
    return val.replace(' -', '')


def split_string_by_integer(s):
    # Find the index of the first occurrence of an integer in the string
    match = re.search(r'\d', s)
    if match:
        split_index = match.start()
        part1 = s[:split_index].strip()
        part2 = s[split_index:].strip()
        part3 = ''
        # Check if there is any text after the integer
        if len(part2) > 0:
            # Find the index of the first occurrence of a non-integer character
            match = re.search(r'\D', part2)
            if match:
                split_index = match.start()
                part3 = part2[split_index:].strip()
                part2 = part2[:split_index].strip()
        return part1, part2, part3
    else:
        return s.strip(), '', ''


# Example usage:
# input_string = "Картопля - 1 кг"
# key, value, unit = split_string_by_integer(input_string)
# print("Part 1:", strip_key_val(key))
# print("Part 2:", value)
# print("Part 3:", unit)

def get_downloads_folder():
    # Get the user's home directory
    home_dir = os.path.expanduser("~")
    # Combine it with the Downloads folder path
    downloads_folder = os.path.join(home_dir, "Downloads")
    return downloads_folder


def convert_units(value, input_unit, desired_unit):
    # Conversion factors for supported units
    supported_units = {"кг": 1, "шт": 1, "г": 1e-3}

    if input_unit not in supported_units or desired_unit not in supported_units:
        return None

    conversion_factor = supported_units[input_unit] / \
        supported_units[desired_unit]

    converted_value = float(value) * conversion_factor
    return converted_value


class ExcelModifier:
    def __init__(self, filename, sheetname=sheet_name):
        self.filename = filename
        # workbook = openpyxl.load_workbook(filename, read_only=True, data_only=True, keep_vba=False)
        self.workbook = openpyxl.load_workbook(filename)
        self.sheetname = sheetname

    def close(self):
        # Close the workbook
        self.workbook.close()

    def find_row_by_key(self, column_index, key):
        # Get the specific sheet
        sheet = self.workbook[self.sheetname]

        # Search for the key in the first column (assuming the key is in the first column)
        for row in sheet.iter_rows():
            print(row[column_index].value)
            if row[column_index].value == key:
                return row[0].row

        # If the key is not found, return None
        return None

    def save(self, output_filename):
        # Save the modified workbook to a new XLSX file
        self.workbook.save(output_filename)

    def set_row_key_value(self, key, value, unit):
        row = self.find_row_by_key(key_name_coulmn_index, key)

        if row is None:
            warnings.warn(f"Key '{key}' not found in '{self.sheetname}'.")

        # Get the specific sheet
        sheet = self.workbook[self.sheetname]
        dest_unit = sheet.cell(row, unit_coulmn_index + 1).value
        new_value = convert_units(value, unit, dest_unit)
        if new_value:
            warnings.warn(
                f"Key '{key}' can not convert units from '{unit}' to '{dest_unit}'.")
        sheet.cell(row, order_coulmn_index + 1, new_value)

# Example usage:


# Setup logging with log.txt as the log file
# setup_logging("log.txt")

# downloads_folder = get_downloads_folder()
# file_name = "Лейт (1).xlsx"
# file_path = os.path.join(downloads_folder, file_name)
# sheet_name = "Лист1"
# key_to_find = "Картопля"

# now = datetime.datetime.now()
# # Format the date as "YYYY-MM-DD_HH-MM-SS"
# formatted_date = now.strftime("%Y-%m-%d_%H-%M-%S")
# new_file_name = f"{formatted_date}_{file_name}"

# excel_modifier = ExcelModifier(file_path)
# excel_modifier.set_row_key_value(key_to_find, value, unit)
# excel_modifier.save(os.path.join(downloads_folder, new_file_name))

def process_files(file1, file2):
    excel_modifier = ExcelModifier(file1)
    try:
        # Open the file in read mode
        with open(file2, "r", encoding="utf-8") as file:
        # Read each line in the file
         for line in file:
            # Process the line or log it
            logging.debug(line.strip())  # Use logging.info() to log each line
            key, value, part3 = split_string_by_integer(line.strip())
            unit = fix_unit(part3)
            logging.info("tuple: %s", (key, value, unit))
            excel_modifier.set_row_key_value(key, value, unit)
        
    except FileNotFoundError:
        logging.error(f"File not found: {file2}")
    except IOError as e:
        logging.error(f"Error reading the file: {e}")
    
    now = datetime.datetime.now()
    # Format the date as "YYYY-MM-DD_HH-MM-SS"
    formatted_date = now.strftime("%Y-%m-%d_%H-%M-%S")
    new_file_name = f"{formatted_date}_{os.path.basename(file1)}"
    excel_modifier.save(new_file_name)

def main():
    parser = argparse.ArgumentParser(description="Process two files.")
    parser.add_argument("excel", help="Path to the excel file.")
    parser.add_argument("text", help="Path to the text file.")
    args = parser.parse_args()

    setup_logging("log.txt")

    process_files(args.excel, args.text)


if __name__ == "__main__":
    main()
